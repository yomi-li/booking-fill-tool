# -*- coding: utf-8 -*-
"""批量下单模板填充（第二个导出选项）。

基于『批量下单标准模版.xlsx』（sheet: 新版客户货物运输委托书）：
- 每票一个「序号」（同一票的所有产品行共享同一序号，1、2、3…）；
- B 列起产品明细与单票模板一致（FBA 箱号连号 / 品名 / SKU / HS / 材质 / 品牌 /
  箱数 / 净毛重 / 个数 / 单价 / 币种 / 长宽高 / 图片(内嵌 AB 列) / 平台链接）；
- AD 参考编号：留空；
- AE PO NO：取亚马逊内部编号（ticket.internal）；
- AF 预计到仓时间：取生成文档的时间；
- AG 地址类型：按上传文件文件名判断——含 Amazon→FBA仓，含 Walmart→WAL仓；
- AH 仓库编码：填仓库代码（WAL 仓必须填，FBA 有则填）；
- AI 及之后所有列（收件国家/地址/邮编/州省/城市/公司/收件人/电话/邮箱/指定路线/
  报关方式/清关方式/是否递延/收件税号/备注）：一律保留模板原有内容与公式不动。

模板中 R 列(产品总个数 =Q*N)与 T 列(申报总价 =R*S)为公式，保留不动。
"""
import datetime
import os
import shutil
import tempfile

import openpyxl

import filler

SHEET_NAME = "新版客户货物运输委托书"

# 1-based 列号（与『新版客户货物运输委托书』表头严格对应）
COL = {
    "seq": 1, "fba_box_no": 2, "cn_name": 3, "en_name": 4, "sku": 5,
    "hs_code": 6, "material_cn": 7, "material_en": 8, "brand": 9,
    "brand_type": 10, "model": 11, "purpose": 12, "electric_magnetic": 13,
    "total_cartons": 14, "net_per_ctn": 15, "gross_per_ctn": 16,
    "qty_per_ctn": 17, "unit_price": 19, "currency": 21,
    "po_price": 22, "po_total": 23, "po_currency": 24,
    "length": 25, "width": 26, "height": 27, "img": 28, "product_link": 29,
    "reference_no": 30, "po_no": 31, "eta": 32, "addr_type": 33, "wh_code": 34,
}
# R 列=18（产品总个数公式 =Q*N）、T 列=20（申报总价公式 =R*S）——保留不动
FORMULA_COLS = {18, 20}
# AI 列(35) 及之后：收件信息等全部保留模板原样（含公式）
KEEP_FROM_COL = 35


def _copy_template_local(template_path: str) -> str:
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    try:
        shutil.copy2(template_path, tmp.name)
    except Exception:
        import subprocess
        win_path = template_path.replace("/", "\\")
        subprocess.run(
            ["powershell", "-NoProfile", "-Command",
             f'Copy-Item -LiteralPath "{win_path}" -Destination "{tmp.name}" -Force'],
            check=True, capture_output=True,
        )
    return tmp.name


def detect_addr_type(source_filename):
    """地址类型：文件名含 Amazon → FBA仓；含 Walmart → WAL仓；否则留空。"""
    fn = (source_filename or "").lower()
    if "amazon" in fn:
        return "FBA仓"
    if "walmart" in fn:
        return "WAL仓"
    return ""


def _has_content(it):
    return any(it.get(k) not in (None, "") for k in
               ("sku", "model", "fba_box_no", "cn_name", "en_name"))


def fill_bulk_template(tickets: list, template_path: str, output_path: str) -> str:
    """tickets: [ {warehouse, fba, internal, source_filename, products/items}, ... ]
    每票生成一行或多行（每产品一行，同票共享序号）。"""
    local = _copy_template_local(template_path)
    try:
        wb = openpyxl.load_workbook(local)
        ws = wb[SHEET_NAME]

        # 清空数据区 A-AC（保留 R/T 公式）与 AD-AH（将重写）；AI 之后不动
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=KEEP_FROM_COL - 1):
            for c in row:
                if c.column in FORMULA_COLS:
                    continue
                c.value = None

        now = datetime.datetime.now()
        r = 2
        for seq, t in enumerate(tickets, 1):
            items = t.get("items") or []
            if not items:
                continue
            internal = (t.get("internal") or "").strip()
            wh = (t.get("warehouse") or "").strip()
            addr_type = detect_addr_type(t.get("source_filename") or "")

            for it in items:
                if not _has_content(it):
                    continue
                # A 序号（同票共享）
                ws.cell(row=r, column=COL["seq"], value=seq)
                # B-AA 产品明细
                ws.cell(row=r, column=COL["fba_box_no"], value=it.get("fba_box_no") or None)
                ws.cell(row=r, column=COL["cn_name"], value=it.get("cn_name") or None)
                ws.cell(row=r, column=COL["en_name"], value=it.get("en_name") or None)
                ws.cell(row=r, column=COL["sku"], value=it.get("sku") or None)
                ws.cell(row=r, column=COL["hs_code"], value=it.get("hs_code") or None)
                ws.cell(row=r, column=COL["material_cn"], value=it.get("material_cn") or None)
                ws.cell(row=r, column=COL["material_en"], value=it.get("material_en") or None)
                ws.cell(row=r, column=COL["brand"], value=it.get("brand") or None)
                ws.cell(row=r, column=COL["brand_type"], value=it.get("brand_type") or None)
                ws.cell(row=r, column=COL["model"], value=it.get("model") or None)
                ws.cell(row=r, column=COL["purpose"], value=it.get("purpose") or None)
                ws.cell(row=r, column=COL["electric_magnetic"],
                        value=it.get("electric_magnetic") or "普货")
                ws.cell(row=r, column=COL["total_cartons"], value=it.get("total_cartons") or None)
                ws.cell(row=r, column=COL["net_per_ctn"], value=it.get("net_per_ctn") or None)
                ws.cell(row=r, column=COL["gross_per_ctn"], value=it.get("gross_per_ctn") or None)
                ws.cell(row=r, column=COL["qty_per_ctn"], value=it.get("qty_per_ctn") or None)
                ws.cell(row=r, column=COL["unit_price"], value=it.get("unit_price") or None)
                ws.cell(row=r, column=COL["currency"],
                        value=(it.get("currency") or "USD") or None)
                ws.cell(row=r, column=COL["po_price"], value=it.get("po_price") or None)
                ws.cell(row=r, column=COL["po_total"], value=it.get("po_total") or None)
                ws.cell(row=r, column=COL["po_currency"],
                        value=(it.get("po_currency") or "CNY") or None)
                ws.cell(row=r, column=COL["length"], value=it.get("length") or None)
                ws.cell(row=r, column=COL["width"], value=it.get("width") or None)
                ws.cell(row=r, column=COL["height"], value=it.get("height") or None)
                # AB 图片（内嵌，AB=28 列）
                filler._add_product_image(ws, r, it.get("image_path"),
                                          sku=it.get("sku", ""), col=COL["img"])
                # AC 平台链接
                ws.cell(row=r, column=COL["product_link"], value=it.get("product_link") or None)
                # AD 参考编号：留空
                # AE PO NO：亚马逊内部编号
                ws.cell(row=r, column=COL["po_no"], value=internal or None)
                # AF 预计到仓时间：生成时间
                eta_cell = ws.cell(row=r, column=COL["eta"], value=now)
                if not eta_cell.number_format or "General" in eta_cell.number_format:
                    eta_cell.number_format = "yyyy-mm-dd hh:mm"
                # AG 地址类型（按文件名）
                ws.cell(row=r, column=COL["addr_type"], value=addr_type or None)
                # AH 仓库编码（WAL 必填 / FBA 有则填）
                ws.cell(row=r, column=COL["wh_code"], value=wh or None)
                # AI 及之后：保留原样（含公式）
                r += 1

        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        return output_path
    finally:
        try:
            os.remove(local)
        except Exception:
            pass
