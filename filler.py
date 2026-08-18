# -*- coding: utf-8 -*-
"""把抽取/编辑后的数据填入固定模板并导出。

环境说明：本机 Excel COM（win32com）在当前 Python 3.13 / pywin32 组合下，
Worksheets/Sheets 的索引访问会返回损坏的代理对象（__call__.Range），
且工作簿 Close 被持久模态框阻塞（0x800ac472），无法稳定使用。
因此默认采用 openpyxl 填充。
关于 LOGO：openpyxl 在 load_workbook 时会自动保留模板内已有的浮动图片
（含其原始锚点），save 后不丢失——故模板自带 LOGO 则天然保留；当前桌面版
模板不含 LOGO（images=0），则不显示，无需额外处理。
产品图片则以「内嵌」方式（TwoCellAnchor 钉在 AD 列对应行）写入，
图片随行移动、与该产品行一一对齐，不再像浮动图那样错位。

每调用一次 fill_template(items, ...) 写入一票（多票请多次调用，分别生成文件）。
"""
import os
import io
import shutil
import tempfile

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU

SHEET_NAME = "客户货物运输委托书 "
# 1-based 列号（与模板表头严格对应）
COL = {
    "fba_box_no": 1, "cn_name": 2, "en_name": 3, "sku": 4, "hs_code": 5,
    "material_cn": 6, "material_en": 7, "brand": 8, "brand_type": 9,
    "model": 10, "purpose": 11, "electric_magnetic": 12,
    "total_cartons": 13, "net_per_ctn": 14, "gross_per_ctn": 15,
    "qty_per_ctn": 16, "total_qty": 17, "unit_price": 18,
    "total_price": 19, "currency": 20, "po_price": 21, "po_total": 22,
    "po_currency": 23, "length": 24, "width": 25, "height": 26,
    "reference_id": 27, "warehouse_code": 28,
    "product_link": 31,
}
IMG_COL = 30  # AD 列：图片

# 模板 LOGO 由 openpyxl 在 load_workbook 时自动保留（浮动图片 + 原锚点）。


def _img_dirs():
    """候选图片目录：优先 DATA_DIR（持久卷），其次项目目录。"""
    base = os.path.dirname(os.path.abspath(__file__))
    data = os.environ.get("DATA_DIR") or base
    out = []
    for d in (data, base):
        p = os.path.join(d, "sku_images")
        if p not in out:
            out.append(p)
    return out


def _resolve_image(image_path, sku=""):
    """图片路径跨平台兜底解析（解决换机/云端后绝对路径失效导致图片不显示）：
    1) 原路径存在 → 直接用；
    2) 按文件名在 sku_images/ 目录下找（云端/另一台电脑上绝对路径不存在的场景）；
    3) 按 SKU 码在 sku_images/ 目录下找（扩展名自动探测）。
    全部找不到返回 None（调用方跳过该行图片）。"""
    if not image_path:
        return None
    if os.path.exists(image_path):
        return image_path
    bn = os.path.basename(str(image_path).replace("\\", "/"))
    for img_dir in _img_dirs():
        if bn:
            cand = os.path.join(img_dir, bn)
            if os.path.exists(cand):
                return cand
        if sku:
            for ext in (".png", ".jpg", ".jpeg", ".webp"):
                cand = os.path.join(img_dir, f"{sku}{ext}")
                if os.path.exists(cand):
                    return cand
    return None


def _compress_image(src, max_px=240):
    """把源图等比缩放到 max_px 内并以压缩 PNG 输出到内存。
    返回 (BytesIO, w_px, h_px)；Pillow 不可用或处理失败时返回 (None,None,None)，
    调用方回退直接嵌入原图。"""
    try:
        from PIL import Image as PILImage
    except Exception:
        return None, None, None
    try:
        with PILImage.open(src) as im:
            im = im.convert("RGB")
            w, h = im.size
            scale = min(1.0, max_px / max(w, h))
            if scale < 1.0:
                im = im.resize((max(1, int(w * scale)), max(1, int(h * scale))),
                               PILImage.LANCZOS)
            buf = io.BytesIO()
            im.save(buf, format="PNG", optimize=True)
            w, h = im.size
            buf.seek(0)
            return buf, w, h
    except Exception:
        return None, None, None

def _has_content(it):
    return any(it.get(k) not in (None, "") for k in
               ("sku", "model", "fba_box_no", "cn_name", "en_name"))


def _copy_template_local(template_path: str) -> str:
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    try:
        shutil.copy2(template_path, tmp.name)
    except Exception:
        import subprocess
        win_path = template_path.replace("/", "\\")
        subprocess.run(
            ["powershell", "-NoProfile", "-Command",
             f'Copy-Item -LiteralPath "{win_path}" -Destination "{tmp.name}" -Force'],
            check=True, capture_output=True,
        )
    return tmp.name


def _cell_width_px(ws, col):
    """估算列宽(像素)。openpyxl 列宽以「字符数」存储，按 Calibri 11 近似换算：
    px ≈ chars*7 + 5（含网格线）。用作把图片约束在单元格内的依据。"""
    letter = get_column_letter(col)
    dim = ws.column_dimensions.get(letter)
    w_chars = dim.width if (dim and dim.width) else (ws.sheet_format.defaultColWidth or 8.43)
    return int(w_chars * 7 + 5)


def _cell_height_px(ws, row):
    """估算行高(像素)。行高以点(pt)存储，1pt=1/72in、1px=1/96in → px=pt*(4/3)。"""
    dim = ws.row_dimensions.get(row)
    if dim and dim.height:
        return int(dim.height * 4 / 3)
    return int((ws.sheet_format.defaultRowHeight or 15) * 4 / 3)


def _add_product_image(ws, row, image_path, sku="", target=95, col=None):
    """在指定列(col, row 行)写入「内嵌」产品图片：图片完整落在 (col,row) 这一个
    单元格内部，左上角不越过单元格左上边界，从而保证内部系统能按单元格正确匹配
    每个 SKU 的图片。

    关键约束（修复点）：
      - 图片先按 target 等比缩放（兼顾可见性），再约束其宽高不超过单元格实际
        像素尺寸（留安全边距），绝不溢出到右侧/下方其它列；
      - 若单元格比图片小，则撑大该单元格（行高 + 图片列宽）以完整容纳图片；
      - 锚点 from=单元格左上角(内缩边距)，to=同一单元格 + 图片宽高(EMU)，
        图片随行/列移动，与产品行一一对应。

    写入前先用 Pillow 把图片等比缩放到 240px 内并压缩（解决超大原图 11-12MB
    被完整嵌入导致 xlsx 巨大、Excel/WPS 解码失败显示空白的问题）；
    路径失效时按文件名/SKU 码兜底解析；Pillow 不可用或处理失败时回退原图。"""
    src = _resolve_image(image_path, sku)
    if not src:
        return
    try:
        buf, w_px, h_px = _compress_image(src)
        if buf is None:
            img = XLImage(src)
            w_px, h_px = img.width, img.height
        else:
            img = XLImage(buf)
            img.width, img.height = w_px, h_px
        if not w_px or not h_px:
            return
        _col = col if col is not None else IMG_COL

        # 1) 基于 target 的等比显示尺寸（保证可见、不超原图）
        scale = min(target / float(h_px), target / float(w_px), 1.0)
        disp_w = int(w_px * scale)
        disp_h = int(h_px * scale)

        # 2) 水平约束：图片不得越过单元格右边界。
        #    若比图片列宽大，则按列宽等比收缩图片（不撑宽列，避免挤压其它列）。
        MARGIN = 4  # 安全边距(px)，保证图片严格在单元格内、不压线
        INSET = 2   # 图片相对单元格左上角的内缩(px)，保证不越过左上边界
        cell_w = _cell_width_px(ws, _col)
        if disp_w > cell_w - MARGIN:
            k = (cell_w - MARGIN) / float(disp_w)
            disp_w = max(1, int(disp_w * k))
            disp_h = max(1, int(disp_h * k))
        img.width, img.height = disp_w, disp_h

        # 3) 垂直：撑大该行的行高以完整容纳图片（图片始终完整可见，
        #    不收缩到极小）。行高(pt) = (图片高+边距)*0.75，px=pt*(4/3) 反推。
        row_h_pt = (disp_h + MARGIN) * 0.75
        ws.row_dimensions[row].height = max(
            ws.row_dimensions[row].height or 0, row_h_pt + 4)

        # 4) 锚点：图片在单元格内居中
        #    水平居中：colOff = (cell_w - disp_w) / 2
        #    垂直居中：rowOff = (cell_h - disp_h) / 2
        cell_h = ws.row_dimensions[row].height * 4/3 if ws.row_dimensions[row].height else _cell_height_px(ws, row)
        center_x = max(0, int((cell_w - disp_w) / 2))
        center_y = max(0, int((cell_h - disp_h) / 2))
        marker_from = AnchorMarker(col=_col - 1, row=row - 1,
                                   colOff=pixels_to_EMU(center_x),
                                   rowOff=pixels_to_EMU(center_y))
        marker_to = AnchorMarker(col=_col - 1, row=row - 1,
                                 colOff=pixels_to_EMU(center_x + disp_w),
                                 rowOff=pixels_to_EMU(center_y + disp_h))
        img.anchor = TwoCellAnchor(_from=marker_from, to=marker_to)
        ws.add_image(img)
    except Exception as e:
        print(f"[PRODUCT IMG ERR] {e}", flush=True)


def fill_template(items: list, template_path: str, output_path: str, currency_default="USD") -> str:
    local = _copy_template_local(template_path)
    try:
        wb = openpyxl.load_workbook(local)
        ws = wb[SHEET_NAME]

        # 清空数据区（保留表头、样式、其它 sheet）
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=31):
            for c in row:
                c.value = None

        # 1) 写数据矩阵
        r = 2
        for it in items:
            if not _has_content(it):
                continue
            for field, col in COL.items():
                v = it.get(field)
                if field == "currency":
                    v = (v or currency_default)
                if v in (None, ""):
                    v = None
                ws.cell(row=r, column=col, value=v)
            r += 1

        # 2) 产品图片（AD 列，按数据行，内嵌 + 压缩，跨平台路径兜底）
        r = 2
        for it in items:
            if not _has_content(it):
                continue
            _add_product_image(ws, r, it.get("image_path"), sku=it.get("sku", ""))
            r += 1

        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        return output_path
    finally:
        try:
            os.remove(local)
        except Exception:
            pass
