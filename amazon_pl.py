# -*- coding: utf-8 -*-
"""Amazon Packing List 解析：多票/单票结构识别、SKU 匹配、托书行生成。

Packing List 结构约定（来自用户实测文件）：
- 表头行：A 列含“地址”，B 列“型号/SKU”，其后 C..Q 为数量/箱数/尺寸等。
- 多票文件：第一组（B 列型号到下一个“合计”）为工厂出货总数，无地址栏 → 跳过（仅作对比）。
  其后的每一组 = 一票订单。
- 单票文件：仅一组，直接作为一票。
- 每组地址栏下：
    第 1 行 A 列 = 仓库编码(warehouse)        → 托书 AB 列
    第 2 行 A 列 = FBA 单号(fba)              → 用于生成 FBA 箱号 A 列
    第 3 行 A 列 = 亚马逊内部编号(internal)   → 托书 AA 列（空则不填）
- 产品行：B 列为 SKU，C 数量 / D 入箱数 / E 总箱数 / F 单箱净重 / G 单箱毛重 /
         H/I/J 零售长宽高(cm)。
- FBA 箱号：整票内按产品顺序连号（U + 6 位零填充）。
    例：FBA19J8CNVGV，产品1 共4箱 → FBA19J8CNVGVU000001-FBA19J8CNVGVU000004；
        产品2 共4箱 → FBA19J8CNVGVU000005-FBA19J8CNVGVU000008。
"""
import io
import re

import openpyxl

import customer_sku


def _is_number(v):
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    s = str(v).strip().replace(",", "")
    try:
        float(s)
        return True
    except ValueError:
        return False


def _num(v):
    if v is None or v == "":
        return ""
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "")
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except ValueError:
        return s


def _find_header_rows(ws):
    """返回所有表头行号（1-based）。表头特征：A 列含'地址' 或 B 列含'型号'/'SKU'。"""
    headers = []
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        a_s = str(a) if a is not None else ""
        b_s = str(b) if b is not None else ""
        if ("地址" in a_s) or ("型号" in b_s) or ("SKU" in b_s.upper()):
            headers.append(r)
    return headers


def parse_packing_list(data: bytes, filename: str = ""):
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[wb.sheetnames[0]]
    wb.close()

    header_rows = _find_header_rows(ws)
    if not header_rows:
        # 退化：整表作为一组
        header_rows = [1]

    # 按表头切分块：[header_i+1, header_{i+1}-1]
    blocks = []
    for i, hr in enumerate(header_rows):
        start = hr + 1
        end = (header_rows[i + 1] - 1) if i + 1 < len(header_rows) else ws.max_row
        blocks.append((hr, start, end))

    tickets = []
    total_block = None
    for hr, start, end in blocks:
        blk = _parse_block(ws, start, end)
        if blk is None:
            continue
        if blk["is_total"]:
            total_block = blk
        else:
            tickets.append(blk)

    # 与 SKU 库匹配
    for t in tickets:
        for p in t["products"]:
            rec = customer_sku.get_sku(p["sku"])
            p["matched"] = rec is not None
            p["record"] = rec

    return {
        "filename": filename,
        "ticket_count": len(tickets),
        "is_multi": len(tickets) > 1,
        "tickets": tickets,
        "factory_total": total_block,
    }


def _parse_block(ws, start, end):
    """解析一个块（不含表头行）。返回 {warehouse,fba,internal,products,is_total} 或 None。"""
    address_codes = []
    products = []
    for r in range(start, end + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        a_s = str(a).strip() if a is not None else ""
        b_s = str(b).strip() if b is not None else ""

        # 地址栏编码：A 列非空且非表头（即使与“合计”同行也要采集，见 260730 文件）
        if a_s and "地址" not in a_s:
            address_codes.append(a_s)

        if b_s == "合计":
            continue

        # 产品行：B 列为非纯数字的 SKU 文本，且本行有 CTN(C) 或 数量(E) 数值
        if b_s and not _is_number(b):
            ctn = _num(ws.cell(r, 5).value)   # E 总箱数
            qty = _num(ws.cell(r, 3).value)   # C 数量
            if _is_number(ctn) or _is_number(qty):
                products.append({
                    "sku": b_s,
                    "qty": _num(ws.cell(r, 3).value),
                    "units_per_ctn": _num(ws.cell(r, 4).value),
                    "ctn": _num(ws.cell(r, 5).value),
                    "nw_per_ctn": _num(ws.cell(r, 6).value),
                    "gw_per_ctn": _num(ws.cell(r, 7).value),
                    "length": _num(ws.cell(r, 8).value),
                    "width": _num(ws.cell(r, 9).value),
                    "height": _num(ws.cell(r, 10).value),
                })

    if not products:
        return None

    warehouse = address_codes[0] if len(address_codes) >= 1 else ""
    fba = address_codes[1] if len(address_codes) >= 2 else ""
    internal = address_codes[2] if len(address_codes) >= 3 else ""
    is_total = (len(address_codes) == 0)  # 无地址栏 = 工厂总数

    return {
        "warehouse": warehouse,
        "fba": fba,
        "internal": internal,
        "is_total": is_total,
        "products": products,
    }


def merge_record(p):
    """把 SKU 库记录合并进产品。

    规格类字段（海关编码/单箱净重/申报单价/长宽高/产品链接/材质/品牌等）
    以【库中已有值为准】（用户需求：托书携带库中的这些信息），
    仅当库为空时才回退到 Packing List 本票解析值。
    数量类字段（数量/箱数/每箱个数/毛重）以本票 PL 解析值为准。
    """
    rec = p.get("record") or {}

    def lib_first(lib_key, pl_key=None, pl_fallback=True):
        """优先取库值；库空则回退到 PL 值（pl_fallback=True 时）。"""
        v = rec.get(lib_key, "")
        if v in (None, ""):
            v = p.get(pl_key if pl_key is not None else lib_key, "") if pl_fallback else ""
        return v

    return {
        "sku": p["sku"],
        "matched": p.get("matched", False),
        "cn_name": rec.get("cn_name", "") or "",
        "en_name": rec.get("en_name", "") or "",
        "hs_code": rec.get("hs_code", "") or "",
        "material_cn": rec.get("material_cn", "") or "",
        "material_en": rec.get("material_en", "") or "",
        "brand": rec.get("brand", "") or "",
        "brand_type": rec.get("brand_type", "") or "",
        "model": rec.get("model", "") or "",
        "purpose": rec.get("purpose", "") or "",
        "electric_magnetic": rec.get("electric_magnetic", "") or "",
        "qty_per_ctn": p.get("units_per_ctn", ""),
        "total_qty": p.get("qty", ""),
        "total_cartons": p.get("ctn", ""),
        # 单箱净重：库优先，PL 回退
        "net_per_ctn": lib_first("net_per_ctn", "nw_per_ctn"),
        "gross_per_ctn": p.get("gw_per_ctn", ""),
        # 长宽高：库优先，PL 回退（满足需求5：携带库中的长宽高）
        "length": lib_first("length"),
        "width": lib_first("width"),
        "height": lib_first("height"),
        "unit_price": rec.get("unit_price", "") or "",
        "currency": rec.get("currency", "") or "USD",
        "po_price": rec.get("po_price", "") or "",
        "po_currency": rec.get("po_currency", "") or "CNY",
        "image_path": rec.get("image_path", "") or "",
        "product_link": rec.get("product_link", "") or "",
    }


def build_items(ticket: dict):
    """把一票转换为托书行（含 FBA 箱号连号）。ticket 可来自 parse 或 UI 回传。"""
    fba = (ticket.get("fba") or "").strip()
    internal = (ticket.get("internal") or "").strip()
    warehouse = (ticket.get("warehouse") or "").strip()
    products = ticket.get("products", [])

    box_start = 1
    items = []
    for p in products:
        m = merge_record(p) if "record" in p else p
        ctn = _num(m.get("total_cartons") or m.get("ctn") or 0)
        try:
            ctn_n = int(float(ctn))
        except (TypeError, ValueError):
            ctn_n = 0
        if fba and ctn_n > 0:
            bs, be = box_start, box_start + ctn_n - 1
            fba_box_no = f"{fba}U{bs:06d}-{fba}U{be:06d}"
        else:
            fba_box_no = ""
        box_start += ctn_n

        # 申报总价
        up = _num(m.get("unit_price") or 0)
        tq = _num(m.get("total_qty") or 0)
        try:
            total_price = round(float(up) * float(tq), 2) if up not in ("", None) and tq not in ("", None) else ""
        except (TypeError, ValueError):
            total_price = ""

        item = {
            "fba_box_no": fba_box_no,
            "cn_name": m.get("cn_name", ""),
            "en_name": m.get("en_name", ""),
            "sku": m.get("sku", ""),
            "hs_code": m.get("hs_code", ""),
            "material_cn": m.get("material_cn", ""),
            "material_en": m.get("material_en", ""),
            "brand": m.get("brand", ""),
            "brand_type": m.get("brand_type", ""),
            "model": m.get("model", ""),
            "purpose": m.get("purpose", ""),
            "electric_magnetic": m.get("electric_magnetic", "") or "普货",
            "total_cartons": ctn,
            "net_per_ctn": m.get("net_per_ctn", ""),
            "gross_per_ctn": m.get("gross_per_ctn", ""),
            "qty_per_ctn": m.get("qty_per_ctn", ""),
            "total_qty": m.get("total_qty", ""),
            "unit_price": m.get("unit_price", ""),
            "total_price": total_price,
            "currency": m.get("currency", "") or "USD",
            "po_price": m.get("po_price", ""),
            "po_total": "",
            "po_currency": m.get("po_currency", "") or "CNY",
            "length": m.get("length", ""),
            "width": m.get("width", ""),
            "height": m.get("height", ""),
            "reference_id": internal,
            "warehouse_code": warehouse,
            "image_path": m.get("image_path", ""),
            "product_link": m.get("product_link", ""),
        }
        items.append(item)
    return items


if __name__ == "__main__":
    import sys
    for path in sys.argv[1:]:
        with open(path, "rb") as f:
            data = f.read()
        res = parse_packing_list(data, path)
        print(f"\n=== {path} ===")
        print(f"多票={res['is_multi']} 票数={res['ticket_count']}")
        if res["factory_total"]:
            ft = res["factory_total"]
            print("工厂总数:", [(p['sku'], p['ctn']) for p in ft['products']])
        for i, t in enumerate(res["tickets"], 1):
            print(f"\n票#{i} 仓库={t['warehouse']} FBA={t['fba']} 内部={t['internal']}")
            for p in t["products"]:
                print(f"  {p['sku']} 箱数={p['ctn']} 匹配={p['matched']}")
            items = build_items(t)
            for it in items:
                print("   FBA箱号:", it["fba_box_no"])
