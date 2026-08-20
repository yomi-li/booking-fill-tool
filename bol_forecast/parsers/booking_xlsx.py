# -*- coding: utf-8 -*-
"""托书（客户货物运输委托书 .xlsx）解析器。

设计要点：
  1. 表头在第 1 行，按「表头中文名」定位列，绝不硬编码列字母
     （不同客户托书列顺序可能不同）。
  2. 聚合算法（与单证口径一致）：
       总箱数 = Σ 总箱数(M)
       总毛重 = Σ 总箱数(M) × 单箱毛重(O)
       总净重 = Σ 总箱数(M) × 单箱净重(N)
       总体积 = Σ 长(X)×宽(Y)×高(Z)/1e6 × 总箱数(M)
  3. 单字段失败不影响整体，错误进 warnings。
"""
from __future__ import annotations

import logging
import re
from typing import Any

import openpyxl

log = logging.getLogger(__name__)

# 表头关键词 -> 角色。匹配时去掉空白/换行后做「包含」判断。
HEADER_MAP = {
    "总箱数": "ctns",
    "单箱净重": "nw_per",
    "单箱毛重": "gw_per",
    "长": "len_cm",
    "宽": "wid_cm",
    "高": "ht_cm",
    "中文品名": "name_cn",
    "英文品名": "name_en",
    "REFERENCE": "po",          # 亚马逊内部编号 REFERENCE ID（PO）
    "海关编码": "hs_code",
    "带电": "magnetic",
}


def _norm(s: Any) -> str:
    return re.sub(r"\s+", "", str(s or "")).upper()


def _to_float(v: Any) -> float | None:
    try:
        if v is None or v == "":
            return None
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _find_booking_sheet(wb) -> Any:
    """优先选含『中文品名』+『总箱数』表头的 sheet。"""
    for ws in wb.worksheets:
        heads = {_norm(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)}
        if any("中文品名" in h for h in heads) and any("总箱数" in h for h in heads):
            return ws
    # 退而求其次：含『总箱数』即可
    for ws in wb.worksheets:
        heads = {_norm(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)}
        if any("总箱数" in h for h in heads):
            return ws
    return None


def parse_booking_xlsx(path: str) -> dict[str, Any]:
    """返回 {fields: {...}, goods: [...], warnings: []}。"""
    result: dict[str, Any] = {"fields": {}, "goods": [], "warnings": []}
    fields = result["fields"]

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        result["warnings"].append(f"托书打开失败: {e}")
        return result

    ws = _find_booking_sheet(wb)
    if ws is None:
        result["warnings"].append("未找到托书数据表（缺少『总箱数』表头）")
        return result

    # ---- 定位列 ----
    col_of: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = _norm(ws.cell(row=1, column=c).value)
        if not h:
            continue
        for kw, role in HEADER_MAP.items():
            if kw.upper() in h and role not in col_of:
                col_of[role] = c

    missing = [kw for kw, role in HEADER_MAP.items()
               if role in ("ctns", "gw_per", "nw_per", "len_cm", "wid_cm", "ht_cm")
               and role not in col_of]
    if missing:
        result["warnings"].append(f"托书缺少关键列: {missing}")

    def val(r, role):
        ci = col_of.get(role)
        return ws.cell(row=r, column=ci).value if ci else None

    # ---- 聚合 ----
    tot_ctns = 0.0
    tot_gw = 0.0
    tot_nw = 0.0
    tot_cbm = 0.0
    tot_units = 0.0
    po_set = []

    for r in range(2, ws.max_row + 1):
        ctns = _to_float(val(r, "ctns"))
        if ctns is None or ctns <= 0:
            # 遇到空行或无效行，视为数据结束（托书一般连续）
            if all(val(r, role) is None or val(r, role) == ""
                   for role in ("ctns", "name_cn", "name_en")):
                continue
            result["warnings"].append(f"第{r}行总箱数无效，已跳过")
            continue

        gw_per = _to_float(val(r, "gw_per")) or 0.0
        nw_per = _to_float(val(r, "nw_per")) or 0.0
        L = _to_float(val(r, "len_cm")) or 0.0
        W = _to_float(val(r, "wid_cm")) or 0.0
        H = _to_float(val(r, "ht_cm")) or 0.0
        units_per = _to_float(val(r, "magnetic"))  # 占位：单箱个数不在 HEADER_MAP

        tot_ctns += ctns
        tot_gw += ctns * gw_per
        tot_nw += ctns * nw_per
        tot_cbm += L * W * H / 1_000_000.0 * ctns

        name_cn = str(val(r, "name_cn") or "").replace("\n", " ").strip()
        name_en = str(val(r, "name_en") or "").replace("\n", " ").strip()
        po = str(val(r, "po") or "").replace("\n", " ").strip()
        if po and po not in po_set:
            po_set.append(po)

        result["goods"].append({
            "row": r,
            "ctns": ctns,
            "gw_per": gw_per,
            "nw_per": nw_per,
            "cbm": round(L * W * H / 1_000_000.0 * ctns, 6),
            "name_cn": name_cn,
            "name_en": name_en,
            "po": po,
            "hs_code": str(val(r, "hs_code") or "").replace("\n", " ").strip(),
        })

    # ---- 收口 ----
    fields["ctns"] = int(round(tot_ctns))
    fields["gw"] = round(tot_gw, 2)
    fields["nw"] = round(tot_nw, 2)
    fields["cbm"] = round(tot_cbm, 4)
    fields["po_list"] = po_set
    fields["row_count"] = len(result["goods"])
    return result
