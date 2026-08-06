# -*- coding: utf-8 -*-
"""依据『Coletti 产品托书汇总.xlsx』（用户最新产品版本）全面同步 SKU 库。

规则（严格按用户需求）：
- 库中已有且 Coletti 内容未变的 SKU → 保持原样不动（跳过写入）；
- 库中已有但内容已变更的 SKU → 更新为 Coletti 最新内容（保留已上传图片 image_path）；
- 库中不存在的 SKU → 作为新增加入。
结果同时写入 _sku_sync_result.txt 便于查看。
"""
import os
import sys
import json

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
import customer_sku

SRC = r"D:/文档/WXWork/1688855820337270/Cache/File/2026-08/Coletti 产品托书汇总.xlsx"
OUT = os.path.join(BASE, "_sku_sync_result.txt")

FIELD_MAP = [
    ("中文品名", "cn_name"),
    ("英文品名", "en_name"),
    ("SKU码", "sku"),
    ("海关编码", "hs_code"),
    ("材质（中文）", "material_cn"),
    ("材质（英文）", "material_en"),
    ("品牌", "brand"),
    ("品牌类型", "brand_type"),
    ("型号", "model"),
    ("用途", "purpose"),
    ("带电", "electric_magnetic"),
    ("单箱", "net_per_ctn"),
    ("毛重", "gross_per_ctn"),
    ("单箱个数", "qty_per_ctn"),
    ("产品总个数", "total_qty"),
    ("申报单价", "unit_price"),
    ("申报总价", "total_price"),
    ("申报币种", "currency"),
    ("长", "length"),
    ("宽", "width"),
    ("高", "height"),
    ("平台链接", "product_link"),
]


def clean_header(h):
    if h is None:
        return ""
    return str(h).replace("\n", "").replace("\r", "").strip()


def build_local_map(headers):
    cleaned = {clean_header(h): h for h in headers if h is not None}
    mapping = {}
    for raw_key, target in FIELD_MAP:
        if raw_key in cleaned:
            mapping[target] = cleaned[raw_key]
            continue
        for ck, orig in cleaned.items():
            if raw_key in ck:
                mapping[target] = ck
                break
    return mapping


def to_num(v):
    if v is None or v == "":
        return ""
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return str(v)


def build_record(raw, lmap):
    """把 Coletti 一行的原始值转成与库一致的记录（仅字段子集）。"""
    rec = {}
    for target, orig_header in lmap.items():
        if target == "sku":
            continue
        val = raw.get(orig_header, "")
        if target in ("net_per_ctn", "gross_per_ctn", "qty_per_ctn",
                      "total_qty", "unit_price", "total_price",
                      "length", "width", "height"):
            rec[target] = to_num(val)
        else:
            rec[target] = str(val).strip() if val is not None else ""
    rec.setdefault("currency", "USD")
    rec.setdefault("po_currency", "CNY")
    if not rec.get("electric_magnetic"):
        rec["electric_magnetic"] = "普货"
    return rec


def sig(rec):
    """比较用的关键字段子集（排除 image_path 等与内容无关字段）。"""
    keys = ["cn_name", "en_name", "hs_code", "material_cn", "material_en",
            "brand", "brand_type", "model", "purpose", "electric_magnetic",
            "net_per_ctn", "gross_per_ctn", "qty_per_ctn", "total_qty",
            "unit_price", "total_price", "currency", "length", "width",
            "height", "product_link"]
    return json.dumps([rec.get(k, "") for k in keys], ensure_ascii=False)


def main():
    import openpyxl
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    lmap = build_local_map(headers)

    existing = customer_sku.load_library()["skus"]
    seen_in_file = set()
    added, updated, unchanged, skipped_dup, failed = [], [], [], [], []

    for r in range(2, ws.max_row + 1):
        raw = {}
        for c, h in enumerate(headers, 1):
            v = ws.cell(r, c).value
            if v is not None:
                raw[clean_header(h)] = v
        if not raw:
            continue
        sku = None
        if "SKU码" in raw and str(raw["SKU码"]).strip():
            sku = str(raw["SKU码"]).strip()
        if not sku:
            failed.append((r, "无 SKU 码"))
            continue
        if sku in seen_in_file:
            skipped_dup.append(sku)
            continue
        seen_in_file.add(sku)

        new_rec = build_record(raw, lmap)
        if sku in existing:
            old = existing[sku]
            if sig(old) == sig(new_rec):
                unchanged.append(sku)          # 内容未变 → 保持原样
                continue
            # 内容已变更 → 更新，保留已上传图片
            merged = dict(new_rec)
            if old.get("image_path"):
                merged["image_path"] = old["image_path"]
            customer_sku.upsert_sku(sku, merged)
            updated.append(sku)
            continue
        customer_sku.upsert_sku(sku, new_rec)
        added.append(sku)

    wb.close()

    lines = []
    lines.append(f"SKU 库同步完成（来源: {os.path.basename(SRC)}）")
    lines.append(f"  新增 SKU: {len(added)}")
    lines.append(f"  内容已变更并更新: {len(updated)}")
    lines.append(f"  内容未变保持原样: {len(unchanged)}")
    lines.append(f"  跳过(文件内重复): {len(skipped_dup)}")
    lines.append(f"  失败: {len(failed)}")
    if added:
        lines.append("\n新增:")
        for s in added:
            lines.append("  + " + s)
    if updated:
        lines.append("\n已更新:")
        for s in updated:
            lines.append("  ~ " + s)
    if skipped_dup:
        lines.append("\n文件内重复(忽略):")
        for s in skipped_dup:
            lines.append("  x " + s)
    if failed:
        lines.append("\n失败明细:")
        for r, why in failed[:10]:
            lines.append(f"  行{r}: {why}")

    out = "\n".join(lines)
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(out)
    print(out)


if __name__ == "__main__":
    main()
