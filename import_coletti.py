# -*- coding: utf-8 -*-
"""从『Coletti 产品托书汇总.xlsx』导入 SKU 到产品库（JSON 持久化）。
结果同时写入 _import_result.txt 便于查看。
"""
import os
import sys
import json

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
import customer_sku

SRC = r"C:\Users\admin\Desktop\Coletti 产品托书汇总.xlsx"
OUT = os.path.join(BASE, "_import_result.txt")

FIELD_MAP = [
    ("中文品名", "cn_name"),
    ("英文品名", "en_name"),
    ("SKU码", "sku"),
    ("海关编码", "hs_code"),
    ("材质（中文）", "material_cn"),
    ("材质（英文）", "material_en"),
    ("品牌", "brand"),
    ("品牌类型", "brand_type"),
    ("型号", "model"),
    ("用途", "purpose"),
    ("带电", "electric_magnetic"),
    ("单箱", "net_per_ctn"),
    ("毛重", "gross_per_ctn"),
    ("单箱个数", "qty_per_ctn"),
    ("产品总个数", "total_qty"),
    ("申报单价", "unit_price"),
    ("申报总价", "total_price"),
    ("申报币种", "currency"),
    ("长", "length"),
    ("宽", "width"),
    ("高", "height"),
    ("平台链接", "product_link"),
]


def clean_header(h):
    if h is None:
        return ""
    return str(h).replace("\n", "").replace("\r", "").strip()


def build_local_map(headers):
    cleaned = {clean_header(h): h for h in headers if h is not None}
    mapping = {}
    for raw_key, target in FIELD_MAP:
        if raw_key in cleaned:
            mapping[target] = cleaned[raw_key]
            continue
        for ck, orig in cleaned.items():
            if raw_key in ck:
                # 用清洗后的键（与 raw 字典键一致），避免换行符导致取不到值
                mapping[target] = ck
                break
    return mapping


def to_num(v):
    if v is None or v == "":
        return ""
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return str(v)


def main():
    import openpyxl
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    lmap = build_local_map(headers)

    existing = customer_sku.load_library()["skus"]
    seen_in_file = set()
    added, refreshed, skipped_dup, failed = [], [], [], []

    for r in range(2, ws.max_row + 1):
        raw = {}
        for c, h in enumerate(headers, 1):
            v = ws.cell(r, c).value
            if v is not None:
                raw[clean_header(h)] = v
        if not raw:
            continue
        sku = None
        if "SKU码" in raw and str(raw["SKU码"]).strip():
            sku = str(raw["SKU码"]).strip()
        if not sku:
            failed.append((r, "无 SKU 码"))
            continue
        if sku in seen_in_file:
            skipped_dup.append(sku)
            continue
        seen_in_file.add(sku)

        rec = {"sku": sku}
        for target, orig_header in lmap.items():
            if target == "sku":
                continue
            val = raw.get(orig_header, "")
            if target in ("net_per_ctn", "gross_per_ctn", "qty_per_ctn",
                          "total_qty", "unit_price", "total_price",
                          "length", "width", "height"):
                rec[target] = to_num(val)
            else:
                rec[target] = str(val).strip() if val is not None else ""
        rec.setdefault("currency", "USD")
        rec.setdefault("po_currency", "CNY")
        if not rec.get("electric_magnetic"):
            rec["electric_magnetic"] = "普货"

        if sku in existing:
            # 库已存在：以 Coletti 完整数据「刷新」（Coletti 为权威源，覆盖旧的不完整数据），
            # 但保留用户已手动上传的产品图片 image_path。
            old = existing[sku]
            merged = dict(rec)  # 用 Coletti 完整记录覆盖
            merged["sku"] = sku
            if old.get("image_path") and not rec.get("image_path"):
                merged["image_path"] = old["image_path"]
            customer_sku.upsert_sku(sku, merged)
            refreshed.append(sku)
            continue

        customer_sku.upsert_sku(sku, rec)
        added.append(sku)

    wb.close()

    lines = []
    lines.append(f"导入完成：")
    lines.append(f"  新增(全新 SKU): {len(added)}")
    lines.append(f"  已存在并按 Coletti 刷新字段: {len(refreshed)}")
    lines.append(f"  跳过(文件内重复): {len(skipped_dup)}")
    lines.append(f"  失败: {len(failed)}")
    if failed:
        lines.append("  失败明细: " + str(failed[:10]))
    lines.append("\n新增 SKU 列表:")
    for s in added:
        lines.append("  + " + s)
    if refreshed:
        lines.append("\n已刷新字段的 SKU:")
        for s in refreshed:
            lines.append("  ~ " + s)
    if skipped_dup:
        lines.append("\n文件内重复(已忽略):")
        for s in skipped_dup:
            lines.append("  x " + s)

    out = "\n".join(lines)
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(out)
    print(out)


if __name__ == "__main__":
    main()
