# -*- coding: utf-8 -*-
"""可配置规则引擎：从 PDF / XLS / XLSX 抽取货物字段，零 AI 依赖。

支持 4 种抽取模式：
  - table         : 在网格(表格/表)里找表头行，按表头文本映射列
  - text          : 对每行文本套正则，捕获字段
  - text_block    : 跨行块正则(报关单多行)
  - grid_relative : Excel 报关单按单元格相对位置抓取

规则来自 rules.json；报关单通过品名包含或总数量与箱单/发票合并，补齐 HS/品牌。
"""
import io
import re
import pdfplumber
import xlrd
import openpyxl

CJK_RE = re.compile(r"[一-鿿]+")
NUM_RE = re.compile(r"[^0-9.\-]")

# ----------------------------------------------------------------------------
# 基础工具
# ----------------------------------------------------------------------------

def _cell_str(v):
    if v is None:
        return ""
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v)
    return str(v).strip()


def _to_float(s):
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s)
    for ch in ["US$", "$", ",", " ", "KGS", "KG", "CBM", "CBMS", "SETS", "CTNS", "套", "台", "PCS", "PC", "CTN", "千克", "个"]:
        s = s.replace(ch, "")
    s = NUM_RE.sub("", s)
    if s in ("", ".", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _split_cn_en(desc):
    if not desc:
        return "", ""
    cn = "".join(CJK_RE.findall(desc))
    en = re.sub(r"[一-鿿]", "", desc).strip()
    return cn.strip(), en.strip()


def _norm(s):
    return re.sub(r"\s+", "", (s or "").lower())


def _clean_num_str(s):
    if s is None:
        return ""
    s = str(s)
    for ch in ["US$", "$", ",", " ", "KGS", "KG", "CBM", "CBMS", "SETS", "CTNS", "台", "PCS", "PC", "CTN"]:
        s = s.replace(ch, "")
    return s.strip()


# ----------------------------------------------------------------------------
# 读取文档 -> 统一结构 {filename, source, text, tables, sheets}
# ----------------------------------------------------------------------------

def read_doc(filename, data):
    low = filename.lower()
    if low.endswith(".pdf"):
        return _read_pdf(filename, data)
    if low.endswith(".xls"):
        return _read_xls(filename, data)
    if low.endswith(".xlsx"):
        return _read_xlsx(filename, data)
    # 兜底当文本
    return {"filename": filename, "source": "text", "text": data.decode("utf-8", "ignore"),
            "tables": [], "sheets": []}


def _read_pdf(filename, data):
    text_parts = []
    tables = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            text_parts.append(t)
            for tb in page.extract_tables():
                grid = [[_cell_str(c) for c in row] for row in tb]
                tables.append(grid)
    return {"filename": filename, "source": "pdf", "text": "\n".join(text_parts),
            "tables": tables, "sheets": []}


def _read_xls(filename, data):
    wb = xlrd.open_workbook(file_contents=data)
    sheets, texts = [], []
    for sh in wb.sheets():
        grid = [[_cell_str(sh.cell_value(r, c)) for c in range(sh.ncols)] for r in range(sh.nrows)]
        sheets.append(grid)
        texts.append("\n".join(" | ".join(c for c in row if c != "") for row in grid))
    return {"filename": filename, "source": "xls", "text": "\n".join(texts),
            "tables": [], "sheets": sheets}


def _read_xlsx(filename, data):
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    sheets, texts = [], []
    for sh in wb.worksheets:
        grid = [[_cell_str(sh.cell_value(r + 1, c + 1)) for c in range(sh.max_column)] for r in range(sh.max_row)]
        sheets.append(grid)
        texts.append("\n".join(" | ".join(c for c in row if c != "") for row in grid))
    return {"filename": filename, "source": "xlsx", "text": "\n".join(texts),
            "tables": [], "sheets": sheets}


# ----------------------------------------------------------------------------
# 文档类型识别 + 规则匹配
# ----------------------------------------------------------------------------

def detect_doc_type(doc):
    u = (doc["filename"] + "\n" + doc["text"]).upper()
    # 报关单标记需精确：仅"报关单"(带"单")/"商品编号"/"海关出口"判为报关单，
    # 避免"报关箱单""报关发票"被误判；报关单正文常含"发票号"会误触 invoice，故放最前
    if "报关单" in u or "商品编号" in u or "海关出口" in u or "CUSTOMS DECLARATION" in u:
        return "customs"
    if "PACKING" in u or "装箱单" in u or "箱单" in u:
        return "packing"
    if "INVOICE" in u or "发票" in u:
        return "invoice"
    return "other"


def rule_matches(rule, doc):
    m = rule.get("match", {})
    if not m:
        return True
    text = doc["text"].lower()
    fname = doc["filename"].lower()
    if "any_text" in m:
        if not any(tok.lower() in text for tok in m["any_text"]):
            return False
    if "all_text" in m:
        if not all(tok.lower() in text for tok in m["all_text"]):
            return False
    if "any_filename" in m:
        if not any(tok.lower() in fname for tok in m["any_filename"]):
            return False
    return True


def find_rule(doc_type, doc, rules):
    for rule in rules:
        if rule.get("doc_type") != doc_type:
            continue
        if rule_matches(rule, doc):
            return rule
    return None


# ----------------------------------------------------------------------------
# 各模式抽取
# ----------------------------------------------------------------------------

def _new_item():
    return {
        "item_no": "", "cn_name": "", "en_name": "", "model": "", "sku": "",
        "hs_code": "", "brand": "", "brand_type": "", "material_cn": "", "material_en": "",
        "purpose": "", "electric_magnetic": "",
        "total_cartons": None, "net_total": None, "gross_total": None, "total_qty": None,
        "total_cbm": None, "net_per_ctn": None, "gross_per_ctn": None, "qty_per_ctn": None,
        "unit_price": None, "total_price": None, "currency": "",
        "po_price": None, "po_total": None, "po_currency": "",
        "length": None, "width": None, "height": None,
    }


def _apply_field(item, field, spec, raw_val, settings):
    if raw_val in (None, ""):
        return
    if "as" in spec and spec["as"] == "num":
        val = _to_float(raw_val)
        if val is None:
            return
    else:
        val = str(raw_val).strip()
    # 货币归一
    if field == "currency" and val:
        alias = settings.get("currency_aliases", {}).get(val.lower())
        if alias:
            val = alias
    item[field] = val


def _extract_table(doc, rule, settings):
    grids = doc.get("tables") or doc.get("sheets") or []
    header_match = [h.lower() for h in rule.get("header_match", [])]
    rf = rule.get("row_filter", {})
    out = []
    for grid in grids:
        # 定位表头行
        hdr_row = -1
        for ri, row in enumerate(grid):
            joined = " ".join(c for c in row if c != "").lower()
            if all(h in joined for h in header_match):
                hdr_row = ri
                break
        if hdr_row < 0:
            continue
        header = grid[hdr_row]
        # 列映射
        col_map = {}
        for field, spec in rule.get("columns", {}).items():
            toks = [t.lower() for t in spec.get("header", [])]
            for ci, cell in enumerate(header):
                cl = cell.lower()
                if any(t in cl for t in toks):
                    col_map[field] = ci
                    break
        # 数据行
        for row in grid[hdr_row + 1:]:
            joined = " ".join(row)
            if not any(v for v in row):
                continue
            if rf.get("exclude_text") and any(t.lower() in joined.lower() for t in rf["exclude_text"]):
                continue
            if rf.get("min_num_cells"):
                nums = sum(1 for c in row if _to_float(c) is not None)
                if nums < rf["min_num_cells"]:
                    continue
            item = _new_item()
            matched = False
            for field, spec in rule.get("columns", {}).items():
                ci = col_map.get(field)
                if ci is None or ci >= len(row):
                    continue
                raw = row[ci]
                if spec.get("mode") == "cn":
                    cn, _ = _split_cn_en(raw)
                    if cn:
                        item["cn_name"] = cn
                        matched = True
                elif spec.get("mode") == "en":
                    _, en = _split_cn_en(raw)
                    if en:
                        item["en_name"] = en
                        matched = True
                else:
                    if spec.get("parse_num_suffix"):
                        raw = _clean_num_str(raw)
                    _apply_field(item, field, spec, raw, settings)
                    if item.get(field) not in (None, ""):
                        matched = True
            if rule.get("currency_default") and not item["currency"]:
                item["currency"] = rule["currency_default"]
            if matched:
                out.append(item)
    return out


def _extract_text(doc, rule, settings):
    patterns = rule.get("patterns", [])
    out = []
    for pat in patterns:
        rx = re.compile(pat["regex"], re.MULTILINE)
        excl = pat.get("exclude_text", [])
        for line in doc["text"].splitlines():
            if excl and any(t.lower() in line.lower() for t in excl):
                continue
            m = rx.search(line)
            if not m:
                continue
            item = _new_item()
            matched = False
            for field, spec in pat.get("fields", {}).items():
                g = spec.get("group")
                if g is None or g > m.lastindex:
                    continue
                _apply_field(item, field, spec, m.group(g), settings)
                if item.get(field) not in (None, ""):
                    matched = True
            if rule.get("currency_default") and not item["currency"]:
                item["currency"] = rule["currency_default"]
            if matched:
                out.append(item)
    return out


def _extract_text_block(doc, rule, settings):
    flags = re.S if "S" in (rule.get("flags") or "") else 0
    rx = re.compile(rule["regex"], flags)
    out = []
    for m in rx.finditer(doc["text"]):
        item = _new_item()
        matched = False
        for field, spec in rule.get("fields", {}).items():
            g = spec.get("group")
            if g is None or g > m.lastindex:
                continue
            _apply_field(item, field, spec, m.group(g), settings)
            if item.get(field) not in (None, ""):
                matched = True
        if rule.get("currency_default") and not item["currency"]:
            item["currency"] = rule["currency_default"]
        if matched:
            out.append(item)
    return out


def _extract_grid_relative(doc, rule, settings):
    sheets = doc.get("sheets") or []
    fr = rule["find_row"]
    col = fr["col"]
    rx = re.compile(fr.get("regex", r".+"))
    out = []
    for grid in sheets:
        for ri, row in enumerate(grid):
            if col >= len(row):
                continue
            if not rx.search(row[col]):
                continue
            item = _new_item()
            matched = False
            for field, spec in rule.get("fields", {}).items():
                r2 = ri + spec.get("row_offset", 0)
                c2 = spec["col"]
                if r2 >= len(grid) or c2 >= len(grid[r2]):
                    continue
                raw = grid[r2][c2]
                if spec.get("parse") == "pipe_5":
                    parts = raw.split("|")
                    raw = parts[5].strip() if len(parts) > 5 else raw
                _apply_field(item, field, spec, raw, settings)
                if item.get(field) not in (None, ""):
                    matched = True
            if rule.get("currency_default") and not item["currency"]:
                item["currency"] = rule["currency_default"]
            if matched:
                out.append(item)
    return out


def extract_with_rule(doc, rule, settings):
    mode = rule.get("mode")
    if mode == "table":
        return _extract_table(doc, rule, settings)
    if mode == "text":
        return _extract_text(doc, rule, settings)
    if mode == "text_block":
        return _extract_text_block(doc, rule, settings)
    if mode == "grid_relative":
        return _extract_grid_relative(doc, rule, settings)
    return []


# ----------------------------------------------------------------------------
# 合并
# ----------------------------------------------------------------------------

def _qkey(v):
    f = _to_float(v)
    if f is None:
        return ""
    try:
        return str(int(f))
    except Exception:
        return str(f)


def _merge_two(a, b):
    for k in a:
        if a[k] in (None, ""):
            a[k] = b.get(k)
        elif b.get(k) not in (None, ""):
            # 都非空：报关单优先补 HS/品牌/品牌类型/中文品名
            if k in ("hs_code", "brand", "brand_type", "cn_name") and b.get(k):
                a[k] = b[k]
    return a


def merge_docs(doc_items):
    """doc_items: list of (doc_type, [items])；返回合并后的 items。"""
    packing, invoice, customs = [], [], []
    for dtype, items in doc_items:
        if dtype == "packing":
            packing += items
        elif dtype == "invoice":
            invoice += items
        elif dtype == "customs":
            customs += items
        else:
            packing += items

    # 1) 箱单 + 发票 按 (cn,en,model,数量) 合并，避免同名不同数量的行被压成一行
    merged = []
    for it in packing + invoice:
        key = (_norm(it["cn_name"]), _norm(it["en_name"]), _norm(it["model"]), _qkey(it["total_qty"]))
        if key == ("", "", "", ""):
            merged.append(it)
            continue
        hit = None
        for m in merged:
            mk = (_norm(m["cn_name"]), _norm(m["en_name"]), _norm(m["model"]), _qkey(m["total_qty"]))
            if mk == key:
                hit = m
                break
        if hit:
            _merge_two(hit, it)
        else:
            merged.append(it)

    # 2) 报关单 合并进已有项：先按总数量 1:1，否则按品名包含应用到全部聚合行
    def find_by_qty(q):
        if not q:
            return None
        for m in merged:
            if m["total_qty"] not in (None, "") and _qkey(m["total_qty"]) == _qkey(q):
                return m
        return None

    def find_all_by_name(cn, en):
        cn, en = _norm(cn), _norm(en)
        out = []
        for m in merged:
            mcn, men = _norm(m["cn_name"]), _norm(m["en_name"])
            if cn and mcn and (cn in mcn or mcn in cn):
                out.append(m)
            elif en and men and (en in men or men in en):
                out.append(m)
        return out

    for c in customs:
        tgt = find_by_qty(c["total_qty"])
        if tgt:
            _merge_two(tgt, c)
            continue
        targets = find_all_by_name(c["cn_name"], c["en_name"])
        if targets:
            for t in targets:
                _merge_two(t, c)
        else:
            merged.append(c)
    return merged


# ----------------------------------------------------------------------------
# 顶层入口
# ----------------------------------------------------------------------------

def rule_extract(docs, rules):
    """docs: [{filename,data}]；返回 {shipment, items, method}。"""
    settings = rules.get("settings", {})
    rule_list = rules.get("rules", [])
    parsed = []
    for d in docs:
        doc = read_doc(d["filename"], d["data"])
        dtype = detect_doc_type(doc)
        rule = find_rule(dtype, doc, rule_list)
        if not rule:
            continue
        items = extract_with_rule(doc, rule, settings)
        parsed.append((dtype, items))

    items = merge_docs(parsed)
    # 派生单箱口径
    for it in items:
        tc = it["total_cartons"]
        if tc in (None, "", 0):
            continue
        tc = _to_float(tc)
        for tot_f, per_f in [("net_total", "net_per_ctn"),
                             ("gross_total", "gross_per_ctn"),
                             ("total_qty", "qty_per_ctn")]:
            if it.get(tot_f) not in (None, ""):
                v = _to_float(it[tot_f])
                if v is not None:
                    it[per_f] = round(v / tc, 3)

    shipment = _extract_shipment(docs, settings)
    return {"shipment": shipment, "items": items, "method": "rule"}


def _extract_shipment(docs, settings):
    sh = {"invoice_no": "", "pi_no": "", "trade_term": "", "currency": "USD",
          "port_of_loading": "", "port_of_discharge": "",
          "shipper": "", "consignee": "", "total_value": None}
    full = "\n".join(d.get("text", "") for d in [read_doc(x["filename"], x["data"]) for x in docs])
    m = re.search(r"INV\s*NO\.?\s*:?\s*(\S+)", full, re.I)
    if m:
        sh["invoice_no"] = m.group(1)
    m = re.search(r"P/?I\s*NO\.?\s*:?\s*(\S+)", full, re.I)
    if m:
        sh["pi_no"] = m.group(1)
    if re.search(r"\bFOB\b", full, re.I):
        sh["trade_term"] = "FOB"
    return sh
